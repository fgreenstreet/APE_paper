from openpyxl import Workbook


def save_figure_data_to_excel(fig, filename='figure_data.xlsx'):
    # Create a new workbook and remove the default sheet
    wb = Workbook()
    wb.remove(wb.active)

    # Extract data from all axes in the figure
    for i, ax in enumerate(fig.get_axes(), start=1):
        # Create a new sheet for each axis
        sheet = wb.create_sheet(title=f'Axis {i}')
        sheet.append(['Line Label', 'X Data', 'Y Data'])

        for line in ax.get_lines():
            x_data = line.get_xdata()
            y_data = line.get_ydata()
            label = line.get_label()
            for x, y in zip(x_data, y_data):
                sheet.append([label, x, y])

    # Save the workbook
    wb.save(filename)
    print(f"Data has been saved to {filename}")

    return


def save_ax_data_to_excel(ax, filename='figure_data.xlsx'):
    # Create a new workbook and remove the default sheet
    wb = Workbook()
    wb.remove(wb.active)


    sheet = wb.create_sheet(title=f'Axis {0}')
    sheet.append(['Line Label', 'X Data', 'Y Data'])

    for line in ax.get_lines():
        x_data = line.get_xdata()
        y_data = line.get_ydata()
        label = line.get_label()
        for x, y in zip(x_data, y_data):
            sheet.append([label, x, y])

    # Save the workbook
    wb.save(filename)
    print(f"Data has been saved to {filename}")

    return


def save_ax_scatter_data_to_excel(ax, filename='scatter_data.xlsx'):
    """Extracts scatter plot data from a Matplotlib axis and saves it to an Excel file."""

    # Create a new workbook and remove the default sheet
    wb = Workbook()
    wb.remove(wb.active)

    # Create a sheet
    sheet = wb.create_sheet(title='Scatter Data')
    sheet.append(['Scatter Label', 'X Data', 'Y Data'])

    # Iterate over all scatter plots in the axis
    for i, collection in enumerate(ax.collections):  # Scatter plots are stored in ax.collections
        try:
            offsets = collection.get_offsets()  # Get scatter points
            label = collection.get_label() if collection.get_label() != "_nolegend_" else f"Scatter {i + 1}"

            for x, y in offsets:
                sheet.append([label, x, y])

        except Exception as e:
            print(f"Skipping scatter {i} due to error: {e}")

    # Save the workbook
    wb.save(filename)
    print(f"Scatter data has been saved to {filename}")

    return